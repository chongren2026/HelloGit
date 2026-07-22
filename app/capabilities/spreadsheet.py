"""将文献检索与分析结果导出为 Excel。"""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models import Analysis
from app.settings import get_settings


HEADERS = (
    "序号",
    "作者",
    "相关性分数",
    "命中关键词",
    "纳入理由",
    "摘要内容",
    "AI总结",
    "文档链接",
)


def _safe_filename(topic: str) -> str:
    return re.sub(r"[^\w\-]+", "_", topic, flags=re.UNICODE).strip("_") or "research"


def _analysis_summary(analysis: Analysis) -> str:
    return "\n".join(
        [
            f"研究背景：{analysis.background}",
            f"研究方法：{analysis.method}",
            f"主要结论：{analysis.finding}",
            f"研究局限：{analysis.limitations}",
        ]
    )


def write_excel_report(
    topic: str,
    analyses: list[Analysis],
    output_dir: Path | None = None,
) -> Path:
    target_dir = output_dir or get_settings().output_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    path = target_dir / f"{_safe_filename(topic)}_检索结果.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "检索结果"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"

    sheet.append(HEADERS)
    for index, analysis in enumerate(analyses, start=1):
        paper = analysis.paper
        sheet.append(
            [
                index,
                ", ".join(paper.authors),
                paper.openalex_relevance_score,
                ", ".join(paper.matched_terms),
                paper.inclusion_reason or "数据源返回的候选文献",
                paper.abstract,
                _analysis_summary(analysis),
                paper.title if paper.source_url else "无可用链接",
            ]
        )
        if paper.source_url:
            link_cell = sheet.cell(row=index + 1, column=8)
            link_cell.hyperlink = paper.source_url
            link_cell.style = "Hyperlink"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_gray = Side(style="thin", color="D9E2F3")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_gray)

    widths = {"A": 8, "B": 24, "C": 14, "D": 26, "E": 48, "F": 80, "G": 68, "H": 42}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[0].alignment = Alignment(horizontal="center", vertical="top")
        row[2].number_format = "0.00"
        sheet.row_dimensions[row[0].row].height = 200

    sheet.auto_filter.ref = f"A1:H{max(sheet.max_row, 1)}"
    if analyses:
        table = Table(displayName="LiteratureResults", ref=f"A1:H{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    workbook.save(path)
    return path
