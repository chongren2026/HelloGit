from openpyxl import load_workbook

from app.capabilities.analyze import analyze_papers
from app.capabilities.spreadsheet import HEADERS, write_excel_report
from app.providers.mock import MockLiteratureProvider


def test_excel_report_contains_required_columns(tmp_path) -> None:
    papers = MockLiteratureProvider().search("测试主题")
    path = write_excel_report("测试主题", analyze_papers(papers), output_dir=tmp_path)

    workbook = load_workbook(path, read_only=True)
    sheet = workbook["检索结果"]
    assert tuple(cell.value for cell in sheet[1]) == HEADERS
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=8).value == papers[0].title
